"""
Источники:
  - data/sample — CSV
  - data/datasets/ — sample CSV, manifest.json, tsv_abs/, json_api/,
  - logs/dbo_events.jsonl, xml_iso20022/, xlsx_crm/
  - manifest.json: additional_loaders (format csv/tsv, column_map)

Параметры:
  {"reset": true} - очистка таблиц (TRUNCATE stg/dwh фактов)
  {"report_date": "2026-04-01"}  — отчётная дата для витрины
"""

from __future__ import annotations

import csv
import io
import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

import pendulum
import psycopg
from airflow.decorators import dag, task
from airflow.operators.python import get_current_context
from bank_dwh import database_url, datasets_data_dir, sample_data_dir
from bank_dwh.airflow_utils import dag_conf, report_date_for_dag
from psycopg.types.json import Json

_DAGS_DIR = Path(__file__).resolve().parent


def _read_csv(path: Path) -> list[dict[str, Any]]:
    """Читает CSV в Unicode: UTF-8 (с BOM), UTF-16 LE/BE, иначе UTF-8 или cp1251"""
    raw = path.read_bytes()
    if raw.startswith(b"\xff\xfe"):
        text = raw.decode("utf-16-le")
    elif raw.startswith(b"\xfe\xff"):
        text = raw.decode("utf-16-be")
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp1251")
    return list(csv.DictReader(io.StringIO(text, newline="")))


def _read_text_any(path: Path) -> str:
    raw = path.read_bytes()
    if raw.startswith(b"\xff\xfe"):
        return raw.decode("utf-16-le")
    if raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16-be")
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return raw.decode("cp1251")


def _read_delimited(path: Path, delimiter: str) -> list[dict[str, Any]]:
    text = _read_text_any(path)
    return list(csv.DictReader(io.StringIO(text, newline=""), delimiter=delimiter))


def _normalize_manifest_row(table: str, row: dict[str, Any]) -> None:
    """Приведение значений из внешних датасетов к ожиданиям DWH."""
    if table == "stg.abs_clients_raw":
        cid = row.get("source_client_id")
        if cid is not None and cid != "":
            row["source_client_id"] = str(cid).strip()
        ct = row.get("client_type")
        if ct:
            u = str(ct).strip().upper()
            if u in ("BUSINESS", "B2B", "B"):
                row["client_type"] = "BUSINESS"
            else:
                row["client_type"] = "INDIVIDUAL"
        if not row.get("status"):
            row["status"] = "ACTIVE"
    elif table == "stg.abs_transactions_raw":
        for k in ("source_transaction_id", "source_client_id", "source_account_id"):
            v = row.get(k)
            if v is not None and str(v).strip() != "":
                row[k] = str(v).strip()
        if not row.get("channel_code"):
            row["channel_code"] = "ONLINE"
        if not row.get("currency_code"):
            row["currency_code"] = "INR"
        if not row.get("direction"):
            row["direction"] = "D"
        if not row.get("status"):
            row["status"] = "POSTED"
    elif table == "stg.abs_accounts_raw":
        if not row.get("account_number_masked"):
            aid = row.get("source_account_id")
            if aid:
                s = str(aid).strip()
                row["account_number_masked"] = f"****{s[-4:]}" if len(s) >= 4 else f"****{s}"


def _read_rows_for_manifest_file(
    path: Path,
    spec: dict[str, Any],
) -> list[dict[str, Any]]:
    fmt = (spec.get("format") or "csv").lower()
    delimiter = str(spec.get("delimiter") or ",")
    if fmt in ("tsv", "tab"):
        return _read_delimited(path, "\t")
    if fmt == "csv" and delimiter not in (",", ""):
        return _read_delimited(path, delimiter)
    if fmt == "csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported manifest format: {fmt}")


def _insert_stg_rows(
    cur: psycopg.Cursor,
    batch_id: int,
    table: str,
    columns: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    col_list = ", ".join(["batch_id", *columns])
    placeholders = ", ".join(["%s"] * (1 + len(columns)))
    sql = f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})"
    for r in rows:
        values: list[Any] = [batch_id]
        for c in columns:
            v = r.get(c)
            values.append(v if v != "" else None)
        cur.execute(sql, values)


def _load_manifest_extra(
    cur: psycopg.Cursor,
    batch_id: int,
    manifest_dir: Path,
) -> None:
    man_path = manifest_dir / "manifest.json"
    if not man_path.is_file():
        return
    doc = json.loads(man_path.read_text(encoding="utf-8"))
    base = manifest_dir.resolve()
    for spec in doc.get("additional_loaders", []):
        rel = spec.get("file")
        table = spec.get("table")
        cols = spec.get("columns")
        if not rel or not table or not cols:
            continue
        columns = tuple(str(c) for c in cols)
        path = (manifest_dir / rel).resolve()
        try:
            path.relative_to(base)
        except ValueError:
            continue
        if not path.is_file():
            continue
        try:
            raw_rows = _read_rows_for_manifest_file(path, spec)
        except ValueError:
            continue
        if not raw_rows:
            continue
        cmap = spec.get("column_map")
        if cmap:
            norm: list[dict[str, Any]] = []
            for r in raw_rows:
                row_out: dict[str, Any] = {c: None for c in columns}
                for csv_header, stg_col in cmap.items():
                    if stg_col not in row_out:
                        continue
                    v = r.get(csv_header)
                    row_out[stg_col] = v if v not in ("", None) else None
                _normalize_manifest_row(table, row_out)
                norm.append(row_out)
            _insert_stg_rows(cur, batch_id, table, columns, norm)
        else:
            for r in raw_rows:
                _normalize_manifest_row(table, r)
            _insert_stg_rows(cur, batch_id, table, columns, raw_rows)


def _map_txn_channel_code(code: str | None) -> str:
    if not code:
        return "ONLINE"
    c = str(code).strip().upper()
    m = {
        "MOBILE": "MB",
        "API": "API",
        "MOBILE_APP": "MB",
    }
    return m.get(c, c)


def _load_tsv_abs_transactions(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    rows: list[dict[str, Any]] = []
    for r in _read_delimited(path, "\t"):
        txid = (r.get("transaction_id") or "").strip()
        if not txid:
            continue
        amt_s = (r.get("amount") or "").replace(",", ".").strip()
        try:
            val = float(amt_s)
        except ValueError:
            continue
        if val < 0:
            direction = "D"
            amt_abs = str(abs(val))
        else:
            direction = "C"
            amt_abs = str(val)
        d = (r.get("transaction_date") or "")[:10].strip()
        op_ts = f"{d}T12:00:00+00:00" if d else None
        rows.append(
            {
                "source_transaction_id": txid,
                "source_client_id": (r.get("client_id") or "").strip(),
                "source_account_id": (r.get("account_id") or "").strip() or None,
                "source_product_id": None,
                "channel_code": _map_txn_channel_code(r.get("channel")),
                "operation_type_code": (r.get("operation_type") or "").strip() or "PAYMENT",
                "operation_ts": op_ts,
                "amount": amt_abs,
                "currency_code": (r.get("currency") or "INR").strip()[:8],
                "direction": direction,
                "status": "POSTED",
                "merchant_name": (r.get("merchant_category") or "").strip() or None,
                "description": (r.get("counterparty") or "").strip() or None,
            },
        )
    if rows:
        _insert_stg_rows(
            cur,
            batch_id,
            "stg.abs_transactions_raw",
            (
                "source_transaction_id", "source_client_id", "source_account_id", "source_product_id",
                "channel_code", "operation_type_code", "operation_ts", "amount", "currency_code",
                "direction", "status", "merchant_name", "description",
            ),
            rows,
        )


def _load_tsv_product_catalog(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    rows: list[dict[str, Any]] = []
    for r in _read_delimited(path, "\t"):
        pid = (r.get("product_id") or "").strip()
        if not pid:
            continue
        rows.append(
            {
                "product_id": int(pid) if str(pid).isdigit() else 0,
                "product_name": (r.get("product_name") or "").strip() or None,
                "product_group": (r.get("product_group") or "").strip() or None,
                "description": (r.get("description") or "").strip() or None,
                "interest_rate": (r.get("interest_rate") or "").strip() or None,
                "credit_limit": (r.get("credit_limit") or "").strip() or None,
                "active_flag": (r.get("active_flag") or "").strip() or None,
                "source_system": (r.get("source_system") or "").strip() or None,
            },
        )
    if not rows:
        return
    _insert_stg_rows(
        cur,
        batch_id,
        "stg.product_catalog_raw",
        (
            "product_id", "product_name", "product_group", "description", "interest_rate",
            "credit_limit", "active_flag", "source_system",
        ),
        rows,
    )


def _load_dbo_jsonl(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    rows: list[dict[str, Any]] = []
    for line in _read_text_any(path).splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            o = json.loads(line)
        except json.JSONDecodeError:
            continue
        payload = {k: v for k, v in o.items() if k in ("ip_country", "account_id", "source_system")}
        rows.append(
            {
                "source_event_id": o.get("event_id"),
                "source_client_id": o.get("customer_id"),
                "channel_code": (o.get("channel") or "").strip(),
                "event_type_code": (o.get("event_type") or "").strip(),
                "event_ts": o.get("event_time"),
                "device_type": (o.get("device_type") or "").strip() or None,
                "platform": None,
                "session_id": (o.get("session_id") or "").strip() or None,
                "success_flag": "true" if o.get("success") else "false" if o.get("success") is not None else None,
                "payload_json": json.dumps(payload, ensure_ascii=False) if payload else None,
            },
        )
    if not rows:
        return
    _insert_stg_rows(
        cur,
        batch_id,
        "stg.dbo_events_raw",
        (
            "source_event_id", "source_client_id", "channel_code", "event_type_code", "event_ts",
            "device_type", "platform", "session_id", "success_flag", "payload_json",
        ),
        rows,
    )


def _load_open_banking_accounts(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    try:
        doc = json.loads(_read_text_any(path))
    except json.JSONDecodeError:
        return
    accs = (doc.get("Data") or {}).get("Account") or []
    if not accs:
        return
    rows: list[dict[str, Any]] = []
    for a in accs:
        st = (a.get("Status") or "").lower()
        status = "ACTIVE" if st == "enabled" else (a.get("Status") or "ACTIVE").upper()
        oa = a.get("StatusUpdateDateTime")
        oad = (oa or "")[:10] if oa else None
        at = f"{a.get('AccountType', '')} / {a.get('AccountSubType', '')}".strip(" /")
        rows.append(
            {
                "source_account_id": a.get("AccountId"),
                "source_client_id": a.get("CustomerId"),
                "account_number_masked": (a.get("Nickname") or f"****{a.get('AccountId', '')}")[:64],
                "account_type": (at or "ACCOUNT")[:50],
                "currency_code": (a.get("Currency") or "INR")[:8],
                "opened_at": oad,
                "closed_at": None,
                "status": status[:30],
            },
        )
    if rows:
        _insert_stg_rows(
            cur,
            batch_id,
            "stg.abs_accounts_raw",
            (
                "source_account_id", "source_client_id", "account_number_masked", "account_type",
                "currency_code", "opened_at", "closed_at", "status",
            ),
            rows,
        )


def _load_open_banking_transactions(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    try:
        doc = json.loads(_read_text_any(path))
    except json.JSONDecodeError:
        return
    items = (doc.get("Data") or {}).get("Transaction") or []
    if not items:
        return
    row_out: list[dict[str, Any]] = []
    for t in items:
        amt = (t.get("Amount") or {}) if isinstance(t.get("Amount"), dict) else {}
        a_s = (amt.get("Amount") or "0").replace(",", ".")
        ind = (t.get("CreditDebitIndicator") or "Debit").lower()
        direction = "D" if "debit" in ind else "C"
        bcode = t.get("BankTransactionCode")
        if isinstance(bcode, dict):
            op = (bcode.get("Code") or "PAYMENT")[:50]
        else:
            op = "PAYMENT"
        bdt = t.get("BookingDateTime")
        m = (t.get("MerchantDetails") or {}) if isinstance(t.get("MerchantDetails"), dict) else {}
        row_out.append(
            {
                "source_transaction_id": t.get("TransactionId"),
                "source_client_id": t.get("CustomerId"),
                "source_account_id": t.get("AccountId"),
                "source_product_id": None,
                "channel_code": "ONLINE",
                "operation_type_code": op,
                "operation_ts": bdt,
                "amount": a_s,
                "currency_code": (amt.get("Currency") or "INR")[:8],
                "direction": direction,
                "status": (t.get("Status") or "POSTED")[:30],
                "merchant_name": (m.get("MerchantName") or "").strip() or None,
                "description": (t.get("TransactionInformation") or "").strip() or None,
            },
        )
    for r in row_out:
        _normalize_manifest_row("stg.abs_transactions_raw", r)
    if row_out:
        _insert_stg_rows(
            cur,
            batch_id,
            "stg.abs_transactions_raw",
            (
                "source_transaction_id", "source_client_id", "source_account_id", "source_product_id",
                "channel_code", "operation_type_code", "operation_ts", "amount", "currency_code",
                "direction", "status", "merchant_name", "description",
            ),
            row_out,
        )


def _complaint_channel(ch: str | None) -> str:
    m = {
        "web": "WEB",
        "phone": "TELEPHONE",
        "email": "EMAIL",
        "referral": "REFERRAL",
    }
    k = (ch or "").strip().lower()
    return m.get(k, "WEB")


def _complaint_status(st: str | None) -> str:
    s = (st or "").lower()
    if "progress" in s:
        return "OPEN"
    if "closed" in s or "monetary" in s or "non-monetary" in s:
        return "CLOSED"
    return "OPEN"


def _load_consumer_complaints_json(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    try:
        doc = json.loads(_read_text_any(path))
    except json.JSONDecodeError:
        return
    complaints = doc.get("complaints") or []
    rows: list[dict[str, Any]] = []
    for c in complaints:
        cid = (c.get("complaint_id") or "").strip()
        if not cid:
            continue
        d0 = (c.get("date_received") or "")[:10]
        ts0 = f"{d0}T00:00:00Z" if d0 else None
        txt = " ".join(
            filter(
                None,
                [
                    c.get("product"),
                    c.get("issue"),
                    (c.get("narrative") or "")[:500],
                ],
            ),
        )
        rows.append(
            {
                "source_appeal_id": cid,
                "source_client_id": (c.get("customer_id") or "").strip(),
                "channel_code": _complaint_channel(c.get("submitted_via")),
                "appeal_type_code": "CONSUMER",
                "created_ts": ts0,
                "resolved_ts": None,
                "status": _complaint_status(c.get("company_response")),
                "priority": "2",
                "result_text": txt or None,
            },
        )
    if not rows:
        return
    _insert_stg_rows(
        cur,
        batch_id,
        "stg.appeals_raw",
        (
            "source_appeal_id", "source_client_id", "channel_code", "appeal_type_code",
            "created_ts", "resolved_ts", "status", "priority", "result_text",
        ),
        rows,
    )


def _local_tag(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag


def _find_acct_othr_id(acct_el: Any) -> str:
    for el in acct_el.iter():
        if _local_tag(el.tag) != "Othr":
            continue
        for sub in el:
            if _local_tag(sub.tag) == "Id" and (sub.text or "").strip().isdigit():
                return (sub.text or "").strip()
    for el in acct_el.iter():
        if _local_tag(el.tag) == "Id" and (el.text or "").strip() and (el.text or "").strip().isdigit():
            return (el.text or "").strip()
    return ""


def _load_camt_file(
    cur: psycopg.Cursor, batch_id: int, path: Path, acc_to_client: dict[str, str], op_type: str,
) -> None:
    if not path.is_file():
        return
    try:
        root = ET.parse(path).getroot()
    except ET.ParseError:
        return
    rows: list[dict[str, Any]] = []
    for parent in root.iter():
        tname = _local_tag(parent.tag)
        if tname not in ("Stmt", "Ntfctn"):
            continue
        acct_id = ""
        for ch in list(parent):
            if _local_tag(ch.tag) == "Acct":
                acct_id = _find_acct_othr_id(ch)
                break
        if not acct_id:
            continue
        base_client = acc_to_client.get(str(acct_id), "")
        for ntry in parent:
            if _local_tag(ntry.tag) != "Ntry":
                continue
            amt_s = "0"
            ccy = "INR"
            for c in ntry:
                if _local_tag(c.tag) == "Amt":
                    amt_s = (c.text or "0").strip() or "0"
                    ccy = (c.get("Ccy") or "INR")[:8]
                    break
            ind = "DBIT"
            for c in ntry:
                if _local_tag(c.tag) == "CdtDbtInd" and c.text:
                    ind = c.text.strip()
                    break
            direction = "D" if "DBIT" in ind.upper() else "C"
            bdt = ""
            for c in ntry:
                if _local_tag(c.tag) == "BookgDt":
                    for s in c.iter():
                        if _local_tag(s.tag) == "Dt" and s.text:
                            bdt = s.text.strip()[:10]
                            break
            ref = ""
            for c in ntry:
                if _local_tag(c.tag) == "AcctSvcrRef" and c.text:
                    ref = c.text.strip()
                    break
            if not ref:
                continue
            rem = ""
            for t in ntry.iter():
                if _local_tag(t.tag) == "Ustrd" and t.text:
                    rem = t.text.strip()
                    break
            source_client = base_client
            for t in ntry.iter():
                if _local_tag(t.tag) == "Nm" and t.text and "Customer" in t.text:
                    m = re.search(r"Customer (\d+)", t.text)
                    if m:
                        source_client = m.group(1)
            if not source_client:
                source_client = base_client
            if not (source_client and str(source_client) != "0"):
                continue
            op_ts = f"{bdt}T12:00:00+00:00" if bdt else None
            try:
                am = abs(float(amt_s.replace(",", ".")))
            except ValueError:
                continue
            rows.append(
                {
                    "source_transaction_id": f"ISO{op_type}-{ref}"[:100],
                    "source_client_id": str(source_client).strip(),
                    "source_account_id": acct_id,
                    "source_product_id": None,
                    "channel_code": "ONLINE",
                    "operation_type_code": "CAMT_STMT" if op_type == "053" else "CAMT_NTF",
                    "operation_ts": op_ts,
                    "amount": str(am),
                    "currency_code": ccy,
                    "direction": direction,
                    "status": "POSTED",
                    "merchant_name": f"camt{op_type}",
                    "description": (rem or f"camt {op_type}")[:2000],
                },
            )
    if not rows:
        return
    for r in rows:
        _normalize_manifest_row("stg.abs_transactions_raw", r)
    _insert_stg_rows(
        cur,
        batch_id,
        "stg.abs_transactions_raw",
        (
            "source_transaction_id", "source_client_id", "source_account_id", "source_product_id",
            "channel_code", "operation_type_code", "operation_ts", "amount", "currency_code",
            "direction", "status", "merchant_name", "description",
        ),
        rows,
    )


def _account_client_map(data_roots: list[Path], manifest_dir: Path | None) -> dict[str, str]:
    m: dict[str, str] = {}
    paths: list[Path] = []
    for root in data_roots:
        p = root / "tsv_abs" / "abs_accounts_external.tsv"
        if p.is_file():
            paths.append(p)
    if manifest_dir and manifest_dir.is_dir():
        p2 = manifest_dir / "tsv_abs" / "abs_accounts_external.tsv"
        if p2.is_file():
            paths.append(p2)
    for p in paths:
        for r in _read_delimited(p, "\t"):
            a = (r.get("account_id") or "").strip()
            c = (r.get("client_id") or "").strip()
            if a and c:
                m[a] = c
    return m


def _load_xlsx_crm_sheets(
    cur: psycopg.Cursor, batch_id: int, path: Path,
) -> None:
    if not path.is_file():
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    mkt = wb["CRM campaigns"] if "CRM campaigns" in wb.sheetnames else None
    risk = wb["Credit card risk"] if "Credit card risk" in wb.sheetnames else None
    if mkt is not None:
        rows: list[dict[str, Any]] = []
        it = mkt.iter_rows(min_row=2, values_only=True)
        for row in it:
            if not row or row[0] is None:
                continue
            rows.append(
                {
                    "campaign_id": str(row[0] or "")[:32],
                    "customer_id": str(row[1] or "")[:32],
                    "age": str(row[2]) if row[2] is not None else None,
                    "job": str(row[3] or "")[:64] if len(row) > 3 else None,
                    "marital": str(row[4] or "")[:32] if len(row) > 4 else None,
                    "education": str(row[5] or "")[:32] if len(row) > 5 else None,
                    "default_flag": str(row[6] or "")[:16] if len(row) > 6 else None,
                    "housing_loan": str(row[7] or "")[:16] if len(row) > 7 else None,
                    "personal_loan": str(row[8] or "")[:16] if len(row) > 8 else None,
                    "contact_channel": str(row[9] or "")[:32] if len(row) > 9 else None,
                    "last_contact_date": str(row[10] or "")[:32] if len(row) > 10 else None,
                    "campaign_contacts": str(row[11] or "")[:16] if len(row) > 11 else None,
                    "previous_outcome": str(row[12] or "")[:32] if len(row) > 12 else None,
                    "term_deposit_subscribed": str(row[13] or "")[:8] if len(row) > 13 else None,
                },
            )
        if rows:
            _insert_stg_rows(
                cur,
                batch_id,
                "stg.crm_marketing_raw",
                (
                    "campaign_id", "customer_id", "age", "job", "marital", "education", "default_flag",
                    "housing_loan", "personal_loan", "contact_channel", "last_contact_date",
                    "campaign_contacts", "previous_outcome", "term_deposit_subscribed",
                ),
                rows,
            )
    if risk is not None:
        rows2: list[dict[str, Any]] = []
        for row in risk.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            rows2.append(
                {
                    "customer_id": str(row[0] or "")[:32],
                    "limit_balance": str(row[1] or "") if len(row) > 1 else None,
                    "sex": str(row[2] or "")[:8] if len(row) > 2 else None,
                    "education": str(row[3] or "")[:32] if len(row) > 3 else None,
                    "marriage": str(row[4] or "")[:32] if len(row) > 4 else None,
                    "age": str(row[5] or "")[:16] if len(row) > 5 else None,
                    "pay_status_m1": str(row[6] or "")[:8] if len(row) > 6 else None,
                    "pay_status_m2": str(row[7] or "")[:8] if len(row) > 7 else None,
                    "bill_amt_m1": str(row[8] or "") if len(row) > 8 else None,
                    "bill_amt_m2": str(row[9] or "") if len(row) > 9 else None,
                    "pay_amt_m1": str(row[10] or "") if len(row) > 10 else None,
                    "pay_amt_m2": str(row[11] or "") if len(row) > 11 else None,
                    "default_next_month": str(row[12] or "")[:4] if len(row) > 12 else None,
                },
            )
        if rows2:
            _insert_stg_rows(
                cur,
                batch_id,
                "stg.credit_risk_raw",
                (
                    "customer_id", "limit_balance", "sex", "education", "marriage", "age", "pay_status_m1",
                    "pay_status_m2", "bill_amt_m1", "bill_amt_m2", "pay_amt_m1", "pay_amt_m2",
                    "default_next_month",
                ),
                rows2,
            )
    wb.close()


def _load_datasets_extras(
    cur: psycopg.Cursor, batch_id: int, data_roots: list[Path], manifest_dir: Path | None,
) -> None:
    roots = list(data_roots)
    if manifest_dir and manifest_dir.is_dir():
        roots.append(manifest_dir)
    mdir = manifest_dir if manifest_dir and manifest_dir.is_dir() else None
    tsv = (mdir / "tsv_abs") if mdir else None
    if tsv and tsv.is_dir():
        _load_tsv_abs_transactions(cur, batch_id, tsv / "abs_transactions_external.tsv")
        _load_tsv_product_catalog(cur, batch_id, tsv / "abs_products_external.tsv")
    for base in roots:
        p = base / "logs" / "dbo_events.jsonl"
        if p.is_file():
            _load_dbo_jsonl(cur, batch_id, p)
    for base in roots:
        ja = base / "json_api" / "open_banking_accounts.json"
        jt = base / "json_api" / "open_banking_transactions.json"
        jc = base / "json_api" / "consumer_complaints_sample.json"
        if ja.is_file():
            _load_open_banking_accounts(cur, batch_id, ja)
        if jt.is_file():
            _load_open_banking_transactions(cur, batch_id, jt)
        if jc.is_file():
            _load_consumer_complaints_json(cur, batch_id, jc)
    acc_m = _account_client_map(data_roots, mdir)
    for base in roots:
        d = base / "xml_iso20022"
        if d.is_dir():
            p53 = d / "camt_053_statement_sample.xml"
            p54 = d / "camt_054_credit_notification_sample.xml"
            if p53.is_file():
                _load_camt_file(cur, batch_id, p53, acc_m, "053")
            if p54.is_file():
                _load_camt_file(cur, batch_id, p54, acc_m, "054")
    for base in roots:
        x = base / "xlsx_crm" / "crm_credit_sources.xlsx"
        if x.is_file():
            _load_xlsx_crm_sheets(cur, batch_id, x)


def load_all_csv(
    conn: psycopg.Connection,
    batch_id: int,
    data_roots: list[Path],
    manifest_dir: Path | None = None,
) -> None:
    loaders: list[tuple[str, str, tuple[str, ...]]] = [
        ("abs_clients.csv", "stg.abs_clients_raw", (
            "source_client_id", "client_type", "first_name", "last_name", "middle_name",
            "birth_date", "registration_date", "tax_id", "phone", "email", "region_code", "status",
        )),
        ("crm_clients.csv", "stg.crm_clients_raw", (
            "source_client_id", "tax_id", "phone", "email", "region_code", "loyalty_tier",
        )),
        ("abs_accounts.csv", "stg.abs_accounts_raw", (
            "source_account_id", "source_client_id", "account_number_masked", "account_type",
            "currency_code", "opened_at", "closed_at", "status",
        )),
        ("abs_products.csv", "stg.abs_products_raw", (
            "source_product_id", "source_client_id", "source_account_id", "product_type_code",
            "product_name", "currency_code", "opened_at", "closed_at", "status", "amount_limit", "interest_rate",
        )),
        ("abs_transactions.csv", "stg.abs_transactions_raw", (
            "source_transaction_id", "source_client_id", "source_account_id", "source_product_id",
            "channel_code", "operation_type_code", "operation_ts", "amount", "currency_code",
            "direction", "status", "merchant_name", "description",
        )),
        ("dbo_events.csv", "stg.dbo_events_raw", (
            "source_event_id", "source_client_id", "channel_code", "event_type_code", "event_ts",
            "device_type", "platform", "session_id", "success_flag", "payload_json",
        )),
        ("appeals.csv", "stg.appeals_raw", (
            "source_appeal_id", "source_client_id", "channel_code", "appeal_type_code",
            "created_ts", "resolved_ts", "status", "priority", "result_text",
        )),
    ]

    with conn.cursor() as cur:
        for filename, table, columns in loaders:
            merged: list[dict[str, Any]] = []
            for root in data_roots:
                p = root / filename
                if p.is_file():
                    merged.extend(_read_csv(p))
            if merged:
                _insert_stg_rows(cur, batch_id, table, columns, merged)

        if manifest_dir and manifest_dir.is_dir():
            _load_manifest_extra(cur, batch_id, manifest_dir)
        _load_datasets_extras(cur, batch_id, data_roots, manifest_dir)


ABS_SOURCE_ID = 1
CRM_SOURCE_ID = 2
DBO_SOURCE_ID = 3


@dataclass
class IdAllocator:
    conn: psycopg.Connection

    def next(self, sql: str) -> int:
        with self.conn.cursor() as cur:
            cur.execute(sql)
            row = cur.fetchone()
            return int(row[0]) + 1 if row and row[0] is not None else 1


def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    return date.fromisoformat(s[:10])


def _parse_ts(s: str | None) -> datetime | None:
    if not s:
        return None
    s = s.strip()
    if not s:
        return None
    dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _parse_decimal(s: str | None) -> Decimal | None:
    if not s or not str(s).strip():
        return None
    return Decimal(str(s).replace(",", "."))


def _bool_from_str(s: str | None) -> bool | None:
    if s is None or s == "":
        return None
    return str(s).lower() in ("1", "true", "yes", "t")


def transform_batch(conn: psycopg.Connection, batch_id: int) -> None:
    alloc_client = IdAllocator(conn)
    alloc_account = IdAllocator(conn)
    alloc_product = IdAllocator(conn)
    alloc_txn = IdAllocator(conn)
    alloc_event = IdAllocator(conn)
    alloc_appeal = IdAllocator(conn)
    alloc_hist = IdAllocator(conn)

    start_client = alloc_client.next("SELECT COALESCE(MAX(client_id), 0) FROM dwh.clients")
    start_account = alloc_account.next("SELECT COALESCE(MAX(account_id), 0) FROM dwh.accounts")
    start_product = alloc_product.next("SELECT COALESCE(MAX(product_id), 0) FROM dwh.products")
    start_txn = alloc_txn.next("SELECT COALESCE(MAX(transaction_id), 0) FROM dwh.transactions")
    start_event = alloc_event.next("SELECT COALESCE(MAX(digital_event_id), 0) FROM dwh.digital_events")
    start_appeal = alloc_appeal.next("SELECT COALESCE(MAX(appeal_id), 0) FROM dwh.appeals")
    start_hist = alloc_hist.next("SELECT COALESCE(MAX(client_history_id), 0) FROM dwh.client_history")

    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT source_client_id, client_type, first_name, last_name, middle_name,
                   birth_date, registration_date, tax_id, phone, email, region_code, status
            FROM stg.abs_clients_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        abs_rows = cur.fetchall()

        cur.execute(
            """
            SELECT source_client_id, tax_id, phone, email, region_code
            FROM stg.crm_clients_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        crm_rows = cur.fetchall()

    tax_to_client: dict[str, int] = {}
    source_to_client: dict[str, int] = {}
    now = datetime.now(timezone.utc)
    hist_id = start_hist - 1

    with conn.cursor() as cur:
        cid = start_client
        for row in abs_rows:
            (
                source_client_id,
                client_type,
                first_name,
                last_name,
                middle_name,
                birth_date_s,
                registration_date_s,
                tax_id,
                phone,
                email,
                region_code,
                status,
            ) = row
            tax_key = (tax_id or "").strip()
            if tax_key and tax_key in tax_to_client:
                client_id = tax_to_client[tax_key]
                cur.execute(
                    """
                    INSERT INTO dwh.client_source_map (
                        source_system_id, source_client_id, client_id, is_primary, match_rule, matched_at
                    ) VALUES (%s, %s, %s, false, 'ABS_DEDUP_TAX', %s)
                    ON CONFLICT (source_system_id, source_client_id) DO UPDATE SET
                        client_id = EXCLUDED.client_id,
                        match_rule = EXCLUDED.match_rule,
                        matched_at = EXCLUDED.matched_at
                    """,
                    (ABS_SOURCE_ID, source_client_id, client_id, now),
                )
            else:
                client_id = cid
                cid += 1
                unified_key = uuid4()
                if tax_key:
                    tax_to_client[tax_key] = client_id
                birth_date = _parse_date(birth_date_s)
                registration_date = _parse_date(registration_date_s)
                full_name = " ".join(p for p in [last_name, first_name, middle_name] if p).strip() or None
                cur.execute(
                    """
                    INSERT INTO dwh.clients (
                        client_id, unified_client_key, client_type, first_name, last_name, middle_name,
                        full_name, birth_date, registration_date, tax_id, phone, email, region_code,
                        status, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    """,
                    (
                        client_id,
                        unified_key,
                        client_type or "INDIVIDUAL",
                        first_name,
                        last_name,
                        middle_name,
                        full_name,
                        birth_date,
                        registration_date,
                        tax_id,
                        phone,
                        email,
                        region_code,
                        status or "ACTIVE",
                        now,
                        now,
                    ),
                )
                hist_id += 1
                cur.execute(
                    """
                    INSERT INTO dwh.client_history (
                        client_history_id, client_id, full_name, phone, email, region_code, status,
                        effective_from, effective_to, is_current
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NULL, true)
                    """,
                    (hist_id, client_id, full_name, phone, email, region_code, status or "ACTIVE", now),
                )
                cur.execute(
                    """
                    INSERT INTO dwh.client_source_map (
                        source_system_id, source_client_id, client_id, is_primary, match_rule, matched_at
                    ) VALUES (%s, %s, %s, true, 'ABS_TAX_KEY', %s)
                    ON CONFLICT (source_system_id, source_client_id) DO UPDATE SET
                        client_id = EXCLUDED.client_id,
                        is_primary = EXCLUDED.is_primary,
                        match_rule = EXCLUDED.match_rule,
                        matched_at = EXCLUDED.matched_at
                    """,
                    (ABS_SOURCE_ID, source_client_id, client_id, now),
                )
            source_to_client[str(source_client_id)] = client_id

        for row in crm_rows:
            source_client_id, tax_id, phone, email, region_code = row
            tax_key = (tax_id or "").strip()
            client_id = tax_to_client.get(tax_key)
            if client_id is None:
                client_id = cid
                cid += 1
                unified_key = uuid4()
                if tax_key:
                    tax_to_client[tax_key] = client_id
                cur.execute(
                    """
                    INSERT INTO dwh.clients (
                        client_id, unified_client_key, client_type, first_name, last_name, middle_name,
                        full_name, birth_date, registration_date, tax_id, phone, email, region_code,
                        status, created_at, updated_at
                    ) VALUES (
                        %s, %s, 'INDIVIDUAL', NULL, NULL, NULL, NULL, NULL, NULL, %s, %s, %s, %s,
                        'ACTIVE', %s, %s
                    )
                    """,
                    (client_id, unified_key, tax_id, phone, email, region_code, now, now),
                )
                hist_id += 1
                cur.execute(
                    """
                    INSERT INTO dwh.client_history (
                        client_history_id, client_id, full_name, phone, email, region_code, status,
                        effective_from, effective_to, is_current
                    ) VALUES (%s, %s, NULL, %s, %s, %s, 'ACTIVE', %s, NULL, true)
                    """,
                    (hist_id, client_id, phone, email, region_code, now),
                )
            cur.execute(
                """
                INSERT INTO dwh.client_source_map (
                    source_system_id, source_client_id, client_id, is_primary, match_rule, matched_at
                ) VALUES (%s, %s, %s, false, 'CRM_TAX_MATCH', %s)
                ON CONFLICT (source_system_id, source_client_id) DO UPDATE SET
                    client_id = EXCLUDED.client_id,
                    matched_at = EXCLUDED.matched_at
                """,
                (CRM_SOURCE_ID, source_client_id, client_id, now),
            )
            source_to_client[f"CRM:{source_client_id}"] = client_id

        cur.execute(
            "SELECT product_type_code, product_type_id FROM dwh.product_types",
        )
        product_type_by_code = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute("SELECT channel_code, channel_id FROM dwh.channels")
        channel_by_code = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute("SELECT operation_type_code, operation_type_id FROM dwh.operation_types")
        op_type_by_code = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute("SELECT event_type_code, event_type_id FROM dwh.event_types")
        event_type_by_code = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute("SELECT appeal_type_code, appeal_type_id FROM dwh.appeal_types")
        appeal_type_by_code = {r[0]: r[1] for r in cur.fetchall()}

        cur.execute(
            """
            SELECT source_account_id, source_client_id, account_number_masked, account_type,
                   currency_code, opened_at, closed_at, status
            FROM stg.abs_accounts_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        account_rows = cur.fetchall()

    account_map: dict[str, int] = {}
    aid = start_account
    with conn.cursor() as cur:
        for row in account_rows:
            (
                source_account_id,
                source_client_id,
                account_number_masked,
                account_type,
                currency_code,
                opened_at_s,
                closed_at_s,
                status,
            ) = row
            client_id = source_to_client.get(str(source_client_id))
            if client_id is None:
                continue
            account_id = aid
            aid += 1
            account_map[str(source_account_id)] = account_id
            cur.execute(
                """
                INSERT INTO dwh.accounts (
                    account_id, source_system_id, source_account_id, client_id, account_number_masked,
                    account_type, currency_code, opened_at, closed_at, status, created_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (source_system_id, source_account_id) DO UPDATE SET
                    client_id = EXCLUDED.client_id,
                    status = EXCLUDED.status,
                    updated_at = EXCLUDED.updated_at
                RETURNING account_id
                """,
                (
                    account_id,
                    ABS_SOURCE_ID,
                    source_account_id,
                    client_id,
                    account_number_masked,
                    account_type,
                    (currency_code or "RUB")[:3],
                    _parse_date(opened_at_s),
                    _parse_date(closed_at_s),
                    status or "ACTIVE",
                    now,
                    now,
                ),
            )
            real_id = cur.fetchone()[0]
            account_map[str(source_account_id)] = real_id

    pid = start_product
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT source_product_id, source_client_id, source_account_id, product_type_code,
                   product_name, currency_code, opened_at, closed_at, status, amount_limit, interest_rate
            FROM stg.abs_products_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        product_rows = cur.fetchall()

    product_map: dict[str, int] = {}
    with conn.cursor() as cur:
        for row in product_rows:
            (
                source_product_id,
                source_client_id,
                source_account_id,
                product_type_code,
                product_name,
                currency_code,
                opened_at_s,
                closed_at_s,
                status,
                amount_limit_s,
                interest_rate_s,
            ) = row
            client_id = source_to_client.get(str(source_client_id))
            if client_id is None:
                continue
            default_pt = next(iter(product_type_by_code.values()), 1)
            ptype_id = product_type_by_code.get((product_type_code or "").strip(), default_pt)
            account_id = account_map.get(str(source_account_id)) if source_account_id else None
            product_id = pid
            pid += 1
            cur.execute(
                """
                INSERT INTO dwh.products (
                    product_id, source_system_id, source_product_id, client_id, account_id, product_type_id,
                    product_name, currency_code, opened_at, closed_at, status, amount_limit, interest_rate,
                    created_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (source_system_id, source_product_id) DO UPDATE SET
                    client_id = EXCLUDED.client_id,
                    account_id = EXCLUDED.account_id,
                    status = EXCLUDED.status,
                    updated_at = EXCLUDED.updated_at
                RETURNING product_id
                """,
                (
                    product_id,
                    ABS_SOURCE_ID,
                    source_product_id,
                    client_id,
                    account_id,
                    ptype_id,
                    product_name,
                    (currency_code or "RUB")[:3] if currency_code else None,
                    _parse_date(opened_at_s),
                    _parse_date(closed_at_s),
                    status or "ACTIVE",
                    _parse_decimal(amount_limit_s),
                    _parse_decimal(interest_rate_s),
                    now,
                    now,
                ),
            )
            real_pid = cur.fetchone()[0]
            product_map[str(source_product_id)] = real_pid

    tid = start_txn
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT source_transaction_id, source_client_id, source_account_id, source_product_id,
                   channel_code, operation_type_code, operation_ts, amount, currency_code,
                   direction, status, merchant_name, description
            FROM stg.abs_transactions_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        txn_rows = cur.fetchall()

        for row in txn_rows:
            (
                source_transaction_id,
                source_client_id,
                source_account_id,
                source_product_id,
                channel_code,
                operation_type_code,
                operation_ts_s,
                amount_s,
                currency_code,
                direction,
                status,
                merchant_name,
                description,
            ) = row
            client_id = source_to_client.get(str(source_client_id))
            if client_id is None:
                continue
            account_id = account_map.get(str(source_account_id)) if source_account_id else None
            product_id = product_map.get(str(source_product_id)) if source_product_id else None
            channel_id = channel_by_code.get(channel_code or "", None)
            op_id = op_type_by_code.get(operation_type_code or "", None)
            op_ts = _parse_ts(operation_ts_s)
            if op_ts is None:
                continue
            op_date = op_ts.date()
            amt = _parse_decimal(amount_s) or Decimal("0")
            transaction_id = tid
            tid += 1
            cur.execute(
                """
                INSERT INTO dwh.transactions (
                    transaction_id, source_system_id, source_transaction_id, client_id, account_id,
                    product_id, channel_id, operation_type_id, operation_ts, operation_date, amount,
                    currency_code, direction, status, merchant_name, description, created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (source_system_id, source_transaction_id) DO UPDATE SET
                    amount = EXCLUDED.amount,
                    status = EXCLUDED.status
                """,
                (
                    transaction_id,
                    ABS_SOURCE_ID,
                    source_transaction_id,
                    client_id,
                    account_id,
                    product_id,
                    channel_id,
                    op_id,
                    op_ts,
                    op_date,
                    amt,
                    (currency_code or "RUB")[:3],
                    direction or "D",
                    status,
                    merchant_name,
                    description,
                    now,
                ),
            )

        cur.execute(
            """
            SELECT source_event_id, source_client_id, channel_code, event_type_code, event_ts,
                   device_type, platform, session_id, success_flag, payload_json
            FROM stg.dbo_events_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        event_rows = cur.fetchall()

        eid = start_event
        for row in event_rows:
            (
                source_event_id,
                source_client_id,
                channel_code,
                event_type_code,
                event_ts_s,
                device_type,
                platform,
                session_id,
                success_flag_s,
                payload_json,
            ) = row
            client_id = source_to_client.get(str(source_client_id))
            if client_id is None:
                continue
            ch_code = (channel_code or "").strip()
            et_code = (event_type_code or "").strip()
            channel_id = channel_by_code.get(ch_code)
            et_id = event_type_by_code.get(et_code)
            if channel_id is None or et_id is None:
                continue
            ev_ts = _parse_ts(event_ts_s)
            if ev_ts is None:
                continue
            try:
                payload_obj: Any = json.loads(payload_json) if payload_json else {}
            except json.JSONDecodeError:
                payload_obj = {}
            cur.execute(
                """
                INSERT INTO dwh.digital_events (
                    digital_event_id, source_system_id, source_event_id, client_id, channel_id, event_type_id,
                    event_ts, event_date, device_type, platform, session_id, success_flag, payload, created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (source_system_id, source_event_id) DO UPDATE SET
                    success_flag = EXCLUDED.success_flag
                """,
                (
                    eid,
                    DBO_SOURCE_ID,
                    source_event_id,
                    client_id,
                    channel_id,
                    et_id,
                    ev_ts,
                    ev_ts.date(),
                    device_type,
                    platform,
                    session_id,
                    _bool_from_str(success_flag_s),
                    Json(payload_obj),
                    now,
                ),
            )
            eid += 1

        apid = start_appeal
        cur.execute(
            """
            SELECT source_appeal_id, source_client_id, channel_code, appeal_type_code,
                   created_ts, resolved_ts, status, priority, result_text
            FROM stg.appeals_raw WHERE batch_id = %s
            ORDER BY id
            """,
            (batch_id,),
        )
        appeal_rows = cur.fetchall()

        for row in appeal_rows:
            (
                source_appeal_id,
                source_client_id,
                channel_code,
                appeal_type_code,
                created_ts_s,
                resolved_ts_s,
                status,
                priority_s,
                result_text,
            ) = row
            client_id = source_to_client.get(str(source_client_id))
            if client_id is None:
                continue
            channel_id = channel_by_code.get(channel_code or "", None)
            at_id = appeal_type_by_code.get(appeal_type_code or "", None)
            c_ts = _parse_ts(created_ts_s)
            if c_ts is None:
                continue
            r_ts = _parse_ts(resolved_ts_s)
            pr: int | None = None
            if priority_s and str(priority_s).strip().isdigit():
                pr = int(priority_s)
            cur.execute(
                """
                INSERT INTO dwh.appeals (
                    appeal_id, source_system_id, source_appeal_id, client_id, channel_id, appeal_type_id,
                    created_ts, resolved_ts, status, priority, result_text, created_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (source_system_id, source_appeal_id) DO UPDATE SET
                    status = EXCLUDED.status,
                    resolved_ts = EXCLUDED.resolved_ts
                """,
                (
                    apid,
                    ABS_SOURCE_ID,
                    source_appeal_id,
                    client_id,
                    channel_id,
                    at_id,
                    c_ts,
                    r_ts,
                    status or "OPEN",
                    pr,
                    result_text,
                    now,
                ),
            )
            apid += 1

        cur.execute(
            """
            SELECT product_id, product_name, product_group, description, interest_rate, credit_limit,
                   active_flag, source_system
            FROM stg.product_catalog_raw WHERE batch_id = %s
            """,
            (batch_id,),
        )
        for pr in cur.fetchall():
            pid, pn, pg, desc, ir, cl, af, ss = pr
            if not pid:
                continue
            flag = (af or "N")[:1].upper() if af else None
            cur.execute(
                """
                INSERT INTO dwh.ref_product_catalog (
                    product_id, product_name, product_group, description, interest_rate, credit_limit,
                    active_flag, source_system
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (product_id) DO UPDATE SET
                    product_name = EXCLUDED.product_name,
                    product_group = EXCLUDED.product_group,
                    description = EXCLUDED.description,
                    interest_rate = EXCLUDED.interest_rate,
                    credit_limit = EXCLUDED.credit_limit,
                    active_flag = EXCLUDED.active_flag,
                    source_system = EXCLUDED.source_system
                """,
                (
                    int(pid),
                    pn,
                    pg,
                    desc,
                    _parse_decimal(str(ir)) if ir else None,
                    _parse_decimal(str(cl)) if cl else None,
                    flag,
                    ss,
                ),
            )

        cur.execute(
            """
            SELECT campaign_id, customer_id, age, job, marital, education, default_flag, housing_loan,
                   personal_loan, contact_channel, last_contact_date, campaign_contacts, previous_outcome,
                   term_deposit_subscribed
            FROM stg.crm_marketing_raw WHERE batch_id = %s
            """,
            (batch_id,),
        )
        for cr in cur.fetchall():
            camp, cust, age_s, job, mar, edu, dfl, hloan, ploan, cch, lcd, ccc, po, tds = cr
            if not camp:
                continue
            cl_id = source_to_client.get(str(cust).strip() if cust else "", None)
            if cl_id is None:
                continue
            age = int(age_s) if age_s and str(age_s).strip().isdigit() else None
            lcd_d = _parse_date(str(lcd)) if lcd else None
            ccn = int(ccc) if ccc and str(ccc).strip().isdigit() else None
            cur.execute(
                """
                INSERT INTO dwh.crm_marketing_leads (
                    client_id, campaign_id, age, job, marital, education, default_flag, housing_loan,
                    personal_loan, contact_channel, last_contact_date, campaign_contacts, previous_outcome,
                    term_deposit_subscribed
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (client_id, campaign_id) DO UPDATE SET
                    age = EXCLUDED.age,
                    job = EXCLUDED.job,
                    term_deposit_subscribed = EXCLUDED.term_deposit_subscribed
                """,
                (
                    cl_id, camp, age, job, mar, edu, dfl, hloan, ploan, cch, lcd_d, ccn, po, tds,
                ),
            )

        cur.execute(
            """
            SELECT customer_id, limit_balance, sex, education, marriage, age, pay_status_m1, pay_status_m2,
                   bill_amt_m1, bill_amt_m2, pay_amt_m1, pay_amt_m2, default_next_month
            FROM stg.credit_risk_raw WHERE batch_id = %s
            """,
            (batch_id,),
        )
        for tr in cur.fetchall():
            cust, lb, sx, edu, mrg, age_s, ps1, ps2, b1, b2, p1, p2, defm = tr
            cl_id = source_to_client.get(str(cust).strip() if cust else "", None)
            if cl_id is None:
                continue
            age = int(age_s) if age_s and str(age_s).strip().isdigit() else None
            try:
                p1i = int(str(ps1).strip()) if ps1 is not None and str(ps1).strip() else None
            except ValueError:
                p1i = None
            try:
                p2i = int(str(ps2).strip()) if ps2 is not None and str(ps2).strip() else None
            except ValueError:
                p2i = None
            dnm = int(defm) if defm is not None and str(defm).strip().isdigit() else None
            cur.execute(
                """
                INSERT INTO dwh.credit_risk_features (
                    client_id, source_customer_id, limit_balance, sex, education, marriage, age,
                    pay_status_m1, pay_status_m2, bill_amt_m1, bill_amt_m2, pay_amt_m1, pay_amt_m2,
                    default_next_month
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (client_id) DO UPDATE SET
                    limit_balance = EXCLUDED.limit_balance,
                    default_next_month = EXCLUDED.default_next_month
                """,
                (
                    cl_id,
                    str(cust).strip(),
                    _parse_decimal(str(lb)) if lb else None,
                    sx,
                    edu,
                    mrg,
                    age,
                    p1i,
                    p2i,
                    _parse_decimal(str(b1)) if b1 else None,
                    _parse_decimal(str(b2)) if b2 else None,
                    _parse_decimal(str(p1)) if p1 else None,
                    _parse_decimal(str(p2)) if p2 else None,
                    dnm,
                ),
            )


def recompute_client_segments(conn: psycopg.Connection, report_date: date) -> None:
    """
    NEW - дата регистрации/создания клиента в пределах 90 дней до отчёта.
    DORMANT - не новый и нет операций/событий 180+ дней (последняя активность по max(транзакция, ДБО)).
    HIGH_VALUE - высокая активность или портфель: оборот 30д >= 250k, или >= 12 операций, или >= 4 активных продукта.
    ACTIVE - умеренная активность 30д: >= 3 операций, или >= 6 цифровых событий, или оборот >= 15k.
    LOW_ACTIVITY - остальные клиенты
    """
    delete_sql = "DELETE FROM dwh.client_segments_history WHERE assigned_by = 'ETL_SEGMENTATION'"
    insert_sql = """
    WITH tx30 AS (
        SELECT
            client_id,
            COUNT(*)::int AS op_30,
            COALESCE(SUM(CASE WHEN direction = 'D' THEN amount ELSE 0 END), 0) AS debit_30,
            COALESCE(SUM(CASE WHEN direction = 'C' THEN amount ELSE 0 END), 0) AS credit_30
        FROM dwh.transactions
        WHERE operation_date >= %(rd)s::date - INTERVAL '30 days'
          AND operation_date <= %(rd)s::date
        GROUP BY client_id
    ),
    last_any AS (
        SELECT client_id, MAX(operation_ts) AS last_tx_ts
        FROM dwh.transactions
        GROUP BY client_id
    ),
    last_dig AS (
        SELECT client_id, MAX(event_ts) AS last_ev_ts
        FROM dwh.digital_events
        GROUP BY client_id
    ),
    de30 AS (
        SELECT client_id, COUNT(*)::int AS ev_30
        FROM dwh.digital_events
        WHERE event_date >= %(rd)s::date - INTERVAL '30 days'
          AND event_date <= %(rd)s::date
        GROUP BY client_id
    ),
    prod_cnt AS (
        SELECT client_id, COUNT(*)::int AS active_products
        FROM dwh.products
        WHERE status = 'ACTIVE'
        GROUP BY client_id
    ),
    metrics AS (
        SELECT
            c.client_id,
            COALESCE(c.registration_date, (c.created_at AT TIME ZONE 'UTC')::date) AS anchor_date,
            COALESCE(t.op_30, 0) AS op_30,
            COALESCE(t.debit_30, 0) + COALESCE(t.credit_30, 0) AS turnover_30,
            la.last_tx_ts,
            ld.last_ev_ts,
            COALESCE(d.ev_30, 0) AS ev_30,
            COALESCE(pc.active_products, 0) AS active_products
        FROM dwh.clients c
        LEFT JOIN tx30 t ON t.client_id = c.client_id
        LEFT JOIN last_any la ON la.client_id = c.client_id
        LEFT JOIN last_dig ld ON ld.client_id = c.client_id
        LEFT JOIN de30 d ON d.client_id = c.client_id
        LEFT JOIN prod_cnt pc ON pc.client_id = c.client_id
    ),
    classified AS (
        SELECT
            client_id,
            CASE
                WHEN anchor_date >= %(rd)s::date - INTERVAL '90 days' THEN 3::smallint
                WHEN anchor_date < %(rd)s::date - INTERVAL '90 days'
                    AND (
                        (last_tx_ts IS NULL AND last_ev_ts IS NULL)
                        OR GREATEST(
                            COALESCE((last_tx_ts AT TIME ZONE 'UTC')::date, DATE '1970-01-01'),
                            COALESCE((last_ev_ts AT TIME ZONE 'UTC')::date, DATE '1970-01-01')
                        ) < %(rd)s::date - INTERVAL '180 days'
                    )
                    THEN 4::smallint
                WHEN turnover_30 >= 250000 OR op_30 >= 12 OR active_products >= 4 THEN 5::smallint
                WHEN op_30 >= 3 OR ev_30 >= 6 OR turnover_30 >= 15000 THEN 1::smallint
                ELSE 2::smallint
            END AS segment_type_id
        FROM metrics
    ),
    mx AS (
        SELECT COALESCE(MAX(segment_history_id), 0) AS base_id
        FROM dwh.client_segments_history
    )
    INSERT INTO dwh.client_segments_history (
        segment_history_id, client_id, segment_type_id, score, effective_from, effective_to,
        is_current, assigned_by
    )
    SELECT
        mx.base_id + ROW_NUMBER() OVER (ORDER BY classified.client_id),
        classified.client_id,
        classified.segment_type_id,
        NULL::numeric,
        %(rd)s::date,
        NULL::date,
        true,
        'ETL_SEGMENTATION'
    FROM classified
    CROSS JOIN mx;
    """
    with conn.cursor() as cur:
        cur.execute(delete_sql)
        cur.execute(insert_sql, {"rd": report_date})


def refresh_profile_export(conn: psycopg.Connection, report_date: date | None = None) -> None:
    report_date = report_date or date.today()
    insert_sql = """
    INSERT INTO dwh.client_profile_export (
        report_date, client_id, unified_client_key, client_type, status, region_code,
        active_accounts_cnt, active_products_cnt,
        debit_turnover_30d, credit_turnover_30d,
        operations_cnt_30d, digital_events_cnt_30d, appeals_cnt_90d,
        last_transaction_ts, last_digital_event_ts, current_segment_type_id
    )
    SELECT
        %(d)s::date,
        c.client_id,
        c.unified_client_key,
        c.client_type,
        c.status,
        c.region_code,
        COALESCE(ac.cnt, 0),
        COALESCE(pc.cnt, 0),
        COALESCE(t.debit_amt, 0),
        COALESCE(t.credit_amt, 0),
        COALESCE(t.op_cnt, 0),
        COALESCE(e.ev_cnt, 0),
        COALESCE(a.ap_cnt, 0),
        t.last_ts,
        e.last_ev_ts,
        s.segment_type_id
    FROM dwh.clients c
    LEFT JOIN LATERAL (
        SELECT COUNT(*)::int AS cnt FROM dwh.accounts a
        WHERE a.client_id = c.client_id AND a.status = 'ACTIVE'
    ) ac ON true
    LEFT JOIN LATERAL (
        SELECT COUNT(*)::int AS cnt FROM dwh.products p
        WHERE p.client_id = c.client_id AND p.status = 'ACTIVE'
    ) pc ON true
    LEFT JOIN LATERAL (
        SELECT
            COALESCE(SUM(CASE WHEN x.direction = 'D' THEN x.amount ELSE 0 END), 0) AS debit_amt,
            COALESCE(SUM(CASE WHEN x.direction = 'C' THEN x.amount ELSE 0 END), 0) AS credit_amt,
            COUNT(*)::int AS op_cnt,
            MAX(x.operation_ts) AS last_ts
        FROM dwh.transactions x
        WHERE x.client_id = c.client_id
          AND x.operation_date >= %(d)s::date - INTERVAL '30 days'
    ) t ON true
    LEFT JOIN LATERAL (
        SELECT COUNT(*)::int AS ev_cnt, MAX(de.event_ts) AS last_ev_ts
        FROM dwh.digital_events de
        WHERE de.client_id = c.client_id
          AND de.event_date >= %(d)s::date - INTERVAL '30 days'
    ) e ON true
    LEFT JOIN LATERAL (
        SELECT COUNT(*)::int AS ap_cnt
        FROM dwh.appeals ap
        WHERE ap.client_id = c.client_id
          AND ap.created_ts::date >= %(d)s::date - INTERVAL '90 days'
    ) a ON true
    LEFT JOIN LATERAL (
        SELECT h.segment_type_id
        FROM dwh.client_segments_history h
        WHERE h.client_id = c.client_id AND h.is_current
        ORDER BY h.effective_from DESC
        LIMIT 1
    ) s ON true
    """
    with conn.cursor() as cur:
        cur.execute(
            "DELETE FROM dwh.client_profile_export WHERE report_date = %s",
            (report_date,),
        )
        cur.execute(insert_sql, {"d": report_date})

def truncate_dwh_and_stg(conn: psycopg.Connection) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            TRUNCATE TABLE
                dwh.client_profile_export,
                dwh.client_segments_history,
                dwh.crm_marketing_leads,
                dwh.credit_risk_features,
                dwh.ref_product_catalog,
                dwh.appeals,
                dwh.digital_events,
                dwh.transactions,
                dwh.products,
                dwh.accounts,
                dwh.client_history,
                dwh.client_source_map,
                dwh.clients
            RESTART IDENTITY CASCADE
            """
        )
        cur.execute(
            """
            TRUNCATE TABLE
                stg.appeals_raw,
                stg.dbo_events_raw,
                stg.credit_risk_raw,
                stg.crm_marketing_raw,
                stg.product_catalog_raw,
                stg.abs_transactions_raw,
                stg.abs_products_raw,
                stg.abs_accounts_raw,
                stg.crm_clients_raw,
                stg.abs_clients_raw,
                stg.load_batch
            RESTART IDENTITY CASCADE
            """
        )


def run_load_sources_to_staging(context: dict[str, Any]) -> int:
    conf = dag_conf(context)
    reset = bool(conf.get("reset", False))
    sample_dir = sample_data_dir(_DAGS_DIR)
    ds_dir = datasets_data_dir(_DAGS_DIR)
    roots = [sample_dir]
    if ds_dir.is_dir():
        roots.append(ds_dir)
    url = database_url()
    with psycopg.connect(url, autocommit=False) as conn:
        if reset:
            truncate_dwh_and_stg(conn)
            conn.commit()

        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO stg.load_batch (source_system_code, status, notes)
                VALUES ('SAMPLE_CSV', 'RUNNING', 'CSV: data/sample + data/datasets')
                RETURNING batch_id
                """
            )
            batch_id = int(cur.fetchone()[0])

        load_all_csv(conn, batch_id, roots, manifest_dir=ds_dir if ds_dir.is_dir() else None)

        with conn.cursor() as cur:
            cur.execute(
                "UPDATE stg.load_batch SET status = %s, finished_at = now() WHERE batch_id = %s",
                ("LOADED", batch_id),
            )
        conn.commit()
    return batch_id


def run_transform_stg_to_dwh(context: dict[str, Any], batch_id: int) -> int:
    url = database_url()
    with psycopg.connect(url, autocommit=False) as conn:
        transform_batch(conn, batch_id)
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE stg.load_batch SET status = %s, notes = notes || ' -> DWH' WHERE batch_id = %s",
                ("DONE", batch_id),
            )
        conn.commit()
    return batch_id


def run_refresh_client_profile_export(context: dict[str, Any]) -> None:
    conf = dag_conf(context)
    report_dt = report_date_for_dag(context, conf)
    url = database_url()
    with psycopg.connect(url, autocommit=False) as conn:
        recompute_client_segments(conn, report_dt)
        refresh_profile_export(conn, report_date=report_dt)
        conn.commit()


@dag(
    dag_id="load_data_to_dwh",
    description="Загрузка из источников (CSV) в stg, преобразование в dwh",
    schedule="@daily",
    start_date=pendulum.datetime(2026, 1, 1, tz="UTC"),
    catchup=False,
    tags=["dwh", "etl", "postgresql"],
    doc_md=__doc__,
    default_args={
        "owner": "dwh",
        "retries": 1,
    },
)
def load_data_to_dwh():
    @task(task_id="load_sources_to_staging")
    def load_sources_to_staging() -> int:
        return run_load_sources_to_staging(get_current_context())

    @task(task_id="transform_stg_to_dwh")
    def transform_stg_to_dwh(batch_id: int) -> int:
        return run_transform_stg_to_dwh(get_current_context(), batch_id)

    @task(task_id="refresh_client_profile_export")
    def refresh_client_profile_export() -> None:
        run_refresh_client_profile_export(get_current_context())

    b = load_sources_to_staging()
    t = transform_stg_to_dwh(b)
    r = refresh_client_profile_export()
    t >> r


load_data_to_dwh()
